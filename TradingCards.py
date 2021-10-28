"""
This assignment is submitted by:
    Name: Shubham Singh
    Student ID: 201538011
"""
# import openpyxl module as opxl
import openpyxl as opxl

# card class
class Card:
    def __init__(self, theName: str, theType: str, theHP: int, isShiny: int, theMoves: list):
        self.cardName: str = theName
        self.cardType: str = theType
        self.cardHP: int   = theHP
        self.cardIsShiny: int = isShiny
        self.cardMoves: list = theMoves
        

    def __str__(self):
        """
        __str__ method to represent 
        class object as string when called
        """
        status = ""
        if self.cardIsShiny == 1:
            status = "shiny"
        elif self.cardIsShiny == 0:
            status = "not shiny"
        return "Card name:" + self.cardName + " of Type: " + self.cardType + " with Health Points: " + str(self.cardHP) + \
               " is " + status + " has Move and Damage Factor tuples as " + '\n'.join(str(item) for item in self.cardMoves)
        

# deck class
class Deck:

    def __init__(self):
        # instantiate the deck class 
        # with an empty list collection
        self.allCards: list = []

    def inputFromFile(self, fileName):
        """
        populate the empty initialised Deck with the
        cards that are present in the given .xlsx file
        """

        # create the workbook object
        wb=opxl.load_workbook(fileName)
        # load the active workbook
        data = wb.active

        # read every row except the first one
        for idx, row in enumerate(data.iter_rows(min_row=2, max_row=data.max_row-1)):
            
            # populate first four columns with card info
            cardName = row[0].value
            cardType = row[1].value
            cardHP = row[2].value
            cardIsShiny = row[3].value

            # add the daamage factor against 
            # move name as a tuple in the list
            cardMoves = []
            # populate the move name and damage factor columns
            for jdx in range(4, data.max_column, 2):
                # if the cells are empty ignore them
                if row[jdx].value == None and row[jdx+1].value == None:
                    continue
                cardMoves.append((row[jdx].value, row[jdx+1].value))
            
            # create the card object
            card = Card(cardName, cardType, cardHP, cardIsShiny, cardMoves)
            # add to the card list
            self.allCards.append(card)

    def __str__(self):
        return "The Deck has total of " + str(len(self.allCards)) + " cards, " + str(self.getShinyCardCount()) + " shiny cards and has average damage value of: " + str(self.getAverageDamage())

    def addCard(self, theCard):
        self.allCards.append(theCard)

    def rmCard(self, theCard):
        """
        only be able to remove a card
        if it is present in the deck
        """
        if theCard in self.allCards:
            self.allCards.remove(theCard)
        else:
            print("This card does not exist in the deck!")

    def getMostPowerful(self):
        """
        return the most powerful card
        with highest average damage factor
        """
        mostPower = float('-inf')
        for idx, card in enumerate(self.allCards):
            list_of_df = []
            for move in card.cardMoves:
                list_of_df.append(move[1])
            avg_dmg = sum(list_of_df)/len(list_of_df)
            if avg_dmg > mostPower:
                mostPower = avg_dmg
                index = idx
        return self.allCards[index]

    def getAverageDamage(self):
        """
        return the average damage factor
        among all cards and their moves
        """
        list_of_avgs = []
        for card in self.allCards:
            list_of_df = []
            for move in card.cardMoves:
                list_of_df.append(move[1])
            avg_dmg = sum(list_of_df)/len(list_of_df)
            list_of_avgs.append(avg_dmg)
        return round(sum(list_of_avgs)/len(list_of_avgs), 1)

    def getShinyCardCount(self):
        """
        return the count of 
        the shiny cards in the deck
        """
        count = 0
        for card in self.allCards:
            if card.cardIsShiny == 1:
                count += 1
        return count

    def viewAllCards(self):
        """
        view information of 
        the cards in the deck
        """
        for index, card in enumerate(self.allCards):
            print("|======================================|")
            print("card", index+1)
            print("Card Name: ", card.cardName)
            print("Card Type: ", card.cardType)
            print("Card Health Points: ", card.cardHP)
            print("Card Shiny Status: ", card.cardIsShiny)
            print("The Move and Damage Factor pairs are: ", *card.cardMoves, sep='\n')

    def viewAllShinyCards(self):
        """
        view information of all the 
        shiny cards in the deck
        """
        for card in self.allCards:
            if card.cardIsShiny == 1:
                print("|======================================|")
                print("Card Name: ", card.cardName)
                print("Card Type: ", card.cardType)
                print("Card Health Points: ", card.cardHP)
                print("Card Shiny Status: ", card.cardIsShiny)
                print("The Move and Damage Factor pairs are: ", *card.cardMoves, sep='\n')

    def viewAllByType(self, theType):
        """
        view information of all the 
        cards in deck by the given type
        """
        for card in self.allCards:
            if card.cardType == theType:
                print("|======================================|")
                print("Card Name: ", card.cardName)
                print("Card Type: ", card.cardType)
                print("Card Health Points: ", card.cardHP)
                print("Card Shiny Status: ", card.cardIsShiny)
                print("The Move and Damage Factor pairs are: ", *card.cardMoves, sep='\n')

    def getCards(self):
        """
        return the card collection 
        in deck as a list object
        """
        return self.allCards

    def saveToFile(self, fileName):
        """
        save the workbook object for
        deck to a given file name 
        """
        wb=opxl.load_workbook(fileName)
        data = wb.active
        
        header = ["Name", "Type","HP", "Shiny", "Move Name 1", "Damage 1","Move Name 2", "Damage 2","Move Name 3", "Damage 3", "Move Name 4", "Damage 4", "Move Name 5", "Damage 5"]
        # populate the first row with the header for card details
        for idx, item in enumerate(header):
            data.cell(row = 1, column = 1+idx).value = item

        # starting second row populate the card information from the deck object
        for idx, card in enumerate(self.allCards):
            
            for row in data.iter_rows(min_row=idx+2, max_row=idx+2, min_col=1, max_col=14):

                row[0].value = card.cardName
                row[1].value = card.cardType
                row[2].value = card.cardHP
                row[3].value = card.cardIsShiny

                jdx = 0
                for idx in range(4, 4+2*len(card.cardMoves), 2):
                    (row[idx].value, row[idx+1].value) = card.cardMoves[jdx]
                    jdx+=1     
        # save the file
        wb.save(fileName)


def main():
    pass
    # deck = Deck()
    # deck.inputFromFile('sampleDeck.xlsx')
    # deck.saveToFile('clearsheet.xlsx')
    # print(deck.viewAllCards())

if __name__ == "__main__":
    main()

